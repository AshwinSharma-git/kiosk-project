from openpyxl import Workbook, load_workbook
import os


def append_to_excel(file_name, data_row, headers=None):
    try:
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            if headers:
                ws.append(headers)
            wb.save(file_name)

        wb = load_workbook(file_name)
        ws = wb.active
        if headers and ws.max_row == 0:
            ws.append(headers)
        ws.append(data_row)
        wb.save(file_name)

    except PermissionError:
        print("⚠️ Cannot write to Excel. Is the file open?")
